import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def analyze_dna_sequence(dna_sequence):
    nucleotide_counts = {'A': 0, 'T': 0, 'G': 0, 'C': 0}
    
    for nucleotide in dna_sequence:
        if nucleotide in nucleotide_counts:
            nucleotide_counts[nucleotide] += 1
    
    total_nucleotides = sum(nucleotide_counts.values())
    nucleotide_frequencies = {nucleotide: count / total_nucleotides for nucleotide, count in nucleotide_counts.items()}
    
    gc_content = (nucleotide_frequencies['G'] + nucleotide_frequencies['C']) * 100
    gc_content = round(gc_content, 2)
    
    motifs = []
    for length in range(2, 6):
        for i in range(len(dna_sequence) - length + 1):
            motif = dna_sequence[i:i + length]
            if dna_sequence.count(motif) > 1 and motif not in motifs:
                motifs.append(motif)
    
    return nucleotide_frequencies, gc_content, motifs

def plot_histogram(nucleotide_frequencies):
    labels = list(nucleotide_frequencies.keys())
    frequencies = list(nucleotide_frequencies.values())

    plt.bar(labels, frequencies)
    plt.xlabel("Nucleotide")
    plt.ylabel("Frequency")
    plt.title("Nucleotide Frequencies")

    plt.savefig("histogram.png")  # Save the chart as an image
    plt.close()  # Close the plot to avoid overlapping figures

def save_to_excel(data, sheet_name, workbook, histogram_image):
    ws = workbook.create_sheet(title=sheet_name)

    for row in data:
        ws.append(row)

    img_cell = ws.cell(row=len(data) + 2, column=1)  # Place the image below the data
    img_cell.value = "Histogram"
    ws.add_image(histogram_image, "B{}".format(len(data) + 3))  # Position the image

def main():
    print("DNA Sequence Analyzer with Excel Export")

    dna_sequence = input("Enter the DNA sequence: ").upper()

    if all(nucleotide in 'ATGC' for nucleotide in dna_sequence):
        nucleotide_frequencies, gc_content, motifs = analyze_dna_sequence(dna_sequence)

        print("Nucleotide Frequencies:")
        for nucleotide, frequency in nucleotide_frequencies.items():
            print(f"{nucleotide}: {frequency:.2f}")

        print(f"GC Content: {gc_content}%")

        if motifs:
            print("Identified DNA Motifs:")
            for motif in motifs:
                print("- " + motif)
        else:
            print("No repeated DNA motifs found.")

        # Create a workbook
        wb = Workbook()

        # Create a histogram
        plot_histogram(nucleotide_frequencies)
        histogram_image = Image("histogram.png")

        # Save analysis results to a worksheet
        data = [
            ["DNA Sequence", "GC Content", "Motifs"],
            [dna_sequence, f"{gc_content}%", "\n".join(motifs)]
        ]
        sheet_name = "Analysis Results"  # Use a fixed sheet name
        save_to_excel(data, sheet_name, wb, histogram_image)

        # Save the workbook with both data and histogram
        excel_file = "dna_sequence_analysis.xlsx"
        wb.save(excel_file)
        print(f"Analysis results saved to '{excel_file}'")

    else:
        print("Invalid DNA sequence entered.")

    print("Thank you for using the DNA Sequence Analyzer!")

if __name__ == "__main__":
    main()
